import tkinter as tk
from tkinter import messagebox
from fpdf import FPDF
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import sys

entry_widgets = {}
FONT_SIZE = 12
paypal_link = "https://paypal.me/JohnathanArcay"
##### Functions #####
def reset_fields():
    for widget in entry_widgets.values():
        widget.delete(0, tk.END)
        
def exit_app():
    if messagebox.askyesno("Exit", "Are you sure you want to exit?"):
        root.destroy()
        sys.exit()

def generate_invoice():
    data = {key: widget.get() for key, widget in entry_widgets.items()}
    
    if not all(data.values()):
        messagebox.showerror("Error", "Please fill in all fields.")
        return
    
    save_location = data['save_location_entry']
    save_location = fr"{save_location}"
    
    try:
        amount_float = float(data['amount_entry'])
    except ValueError:
        messagebox.showerror("Error", "Amount must be a number.")
        return
    
    # Generate PDF Invoice
    try:
        pdf = FPDF()
        pdf.add_page()
        effective_width = pdf.w - 2 * pdf.l_margin
        
        pdf.set_font("helvetica", size=FONT_SIZE)
        pdf.cell(effective_width, 20, text=f'{data['business_entry'].upper()} INVOICE', new_x='LMARGIN', new_y='NEXT', align='C')
        pdf.cell(effective_width, 10, text=f'Date: {data['date_entry']}', new_x='LMARGIN', new_y='NEXT', border=1, align='L')
        pdf.cell(effective_width, 10, text=f'Client Name: {data['client_entry']}', new_x='LMARGIN', new_y='NEXT', border=1, align='L')
        pdf.cell(effective_width, 10, text=f'Amount: ${data['amount_entry']}', new_x='LMARGIN', new_y='NEXT', border=1, align='L')
        pdf.cell(effective_width, 10, text=f'Description: {data['description_entry']}', new_x='LMARGIN', new_y='NEXT', border=1, align='L')
        pdf.ln(10)
        pdf.cell(effective_width, 10, text='Thank you for your business!', new_x='LMARGIN', new_y='NEXT', align='C')
        pdf.set_font("helvetica", style="B", size=FONT_SIZE)
        pdf.cell(effective_width, 10, text=f'Please make payment to {paypal_link}', new_x='LMARGIN', new_y='NEXT', align='C')
        
        invoice_filename = os.path.join(save_location, f'invoice_{data['client_entry']}_{data['business_entry'].upper()}_{datetime.now().strftime("%Y-%m-%d")}.pdf')
        pdf.output(invoice_filename)
    except Exception as e:
        messagebox.showerror("PDF Error", f"Could not save PDF: {e}")
        return
    
    messagebox.showinfo("Success", f"Invoice generated and saved successfully!")
    reset_fields()

##### Main Application #####
root = tk.Tk()
root.title("Invoice Generator")
root.geometry("450x400")

fields = [("Date (YYYY-MM-DD)", "date_entry"),
          ("Client Name", "client_entry"),
          ("Business Name", "business_entry"),
          ("Amount", "amount_entry"),
          ("Description", "description_entry"),
          ("Save Location", "save_location_entry")]

for i, (label_text, var_name) in enumerate(fields):
    tk.Label(root, text=label_text).grid(row=i, column=0, sticky="e", padx=5, pady=5)
    entry = tk.Entry(root, width=30)
    entry.grid(row=i, column=1, padx=5, pady=5)
    entry_widgets[var_name] = entry
    
tk.Button(root, text="Generate Invoice", bg="lightblue", width=15, command=generate_invoice).grid(row=len(fields) + 1, column=0, columnspan=1, pady=20)
tk.Button(root, text="Reset", bg="lightgray", width=10, command=reset_fields).grid(row=len(fields) + 1, column=1, columnspan=1)
tk.Button(root, text="Exit", bg="lightcoral", width=10, command=exit_app).grid(row=len(fields) + 1, column=2, columnspan=1, pady=10)

root.mainloop()